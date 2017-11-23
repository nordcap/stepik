from openpyxl import Workbook
from openpyxl import load_workbook


wb = load_workbook(filename='marka.xlsx')

ws = wb['sum marka']

out = dict()

all_list = [(ws['A' + str(i)].value, ws['B' + str(i)].value) for i in range(1, 6)]

for cell in all_list:
    if cell[0] in out:
        out[cell[0]] += cell[1]
    else:
        out[cell[0]] = cell[1]

wb1 = Workbook()
ws1 = wb1.active
ws1.title = 'sum'
row = 1
for key in sorted(out):
    ws1.cell(column=1, row=row, value = key)
    ws1.cell(column=2, row=row, value = out[key])
    row += 1

wb1.save(filename='new_marka.xlsx')
